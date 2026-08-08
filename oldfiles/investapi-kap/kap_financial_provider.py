import json
import logging
import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import requests
from sqlalchemy.orm import Session

from database import Company
from models.financial_statements import (
    CurrencyEnum,
    FinancialAccount,
    FinancialAccountSourceMap,
    FinancialPeriod,
    FinancialValue,
    PeriodTypeEnum,
    ReportTypeEnum,
    StatementTypeEnum,
)

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class KapItemSpec:
    item_id: str
    website_label: str


class KapFinancialProvider:
    """Fetch summary financial statement items from KAP via compare-items + export/compareItems.

    This provider intentionally uses the "Finansal Tablo Kalem Sorgulama" API surface, because
    some company-detail endpoints are protected (HTTP 666).

    Result is a flat summary table (not full hierarchical statement). We still store it into
    financial_statements tables, and can later reconcile/drill-down using other providers.
    """

    BASE_URL = "https://www.kap.org.tr"

    SEARCH_COMBINED_URL = f"{BASE_URL}/tr/api/search/combined"
    COMPANY_ITEMS_URL_TEMPLATE = f"{BASE_URL}/tr/api/company/items/ALL/A"

    COMPARE_ITEMS_BY_SECTOR_URL_TEMPLATE = (
        f"{BASE_URL}/tr/api/analysis/compare-items-by-sector/{{sector}}"
    )
    EXPORT_COMPARE_ITEMS_URL = f"{BASE_URL}/tr/api/export/compareItems"

    USER_AGENT = (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0 Safari/537.36"
    )

    def __init__(self, sector: str = "GENERAL"):
        self.sector = sector

    def _headers(self) -> Dict[str, str]:
        return {
            "User-Agent": self.USER_AGENT,
            "Accept": "application/json,text/plain,*/*",
            "Content-Type": "application/json;charset=UTF-8",
        }

    def _get_json(self, url: str) -> Any:
        resp = requests.get(url, headers={"User-Agent": self.USER_AGENT, "Accept": "application/json"}, timeout=30)
        resp.raise_for_status()
        return resp.json()

    def _post_json(self, url: str, payload: Dict[str, Any]) -> Any:
        resp = requests.post(url, headers=self._headers(), data=json.dumps(payload).encode("utf-8"), timeout=30)
        resp.raise_for_status()
        return resp.json()

    def _post_bytes(self, url: str, payload: Dict[str, Any]) -> Tuple[bytes, str]:
        resp = requests.post(
            url,
            headers={
                "User-Agent": self.USER_AGENT,
                "Accept": "*/*",
                "Content-Type": "application/json;charset=UTF-8",
            },
            data=json.dumps(payload).encode("utf-8"),
            timeout=60,
        )
        resp.raise_for_status()
        return resp.content, resp.headers.get("content-type", "")

    def _resolve_kap_member_title(self, mkk_member_oid: str) -> Optional[str]:
        """Resolve the exact KAP company title used by export/compareItems.

        KAP export endpoint seems to require kapMemberTitle as seen in
        /tr/api/company/items/ALL/A.
        """
        try:
            companies = self._get_json(self.COMPANY_ITEMS_URL_TEMPLATE)
            for c in companies or []:
                if str(c.get("mkkMemberOid") or "").strip() == str(mkk_member_oid).strip():
                    title = c.get("kapMemberTitle")
                    return str(title).strip() if title else None
        except Exception:
            logger.exception("Failed to resolve kapMemberTitle for %s", mkk_member_oid)
        return None

    def _resolve_company_type(self, mkk_member_oid: str) -> Optional[str]:
        try:
            companies = self._get_json(self.COMPANY_ITEMS_URL_TEMPLATE)
            for c in companies or []:
                if str(c.get("mkkMemberOid") or "").strip() == str(mkk_member_oid).strip():
                    ct = c.get("kapMemberType")
                    return str(ct).strip() if ct else None
        except Exception:
            logger.exception("Failed to resolve kapMemberType for %s", mkk_member_oid)
        return None

    def _find_member_oid_title_and_type(self, ticker: str) -> Tuple[str, str, str]:
        data = self._post_json(self.SEARCH_COMBINED_URL, {"keyword": ticker.upper()})
        for block in data:
            if block.get("category") != "companyOrFunds":
                continue
            for r in block.get("results", []) or []:
                if (r.get("cmpOrFundCode") or "").lower() == ticker.lower():
                    member_oid = r.get("memberOrFundOid")
                    if not member_oid:
                        continue
                    # Prefer exact KAP title from company items list
                    title = self._resolve_kap_member_title(str(member_oid))
                    if not title:
                        title = r.get("searchValue")
                    company_type = self._resolve_company_type(str(member_oid))
                    if not company_type:
                        company_type = str(r.get("kapMemberType") or "").strip() or "IGS"
                    if title and company_type:
                        return str(member_oid), str(title), str(company_type)
        raise ValueError(f"KAP search could not resolve ticker {ticker}")

    def _get_compare_items(self) -> List[KapItemSpec]:
        url = self.COMPARE_ITEMS_BY_SECTOR_URL_TEMPLATE.format(sector=self.sector)
        items = self._get_json(url)
        specs: List[KapItemSpec] = []
        for it in items or []:
            iid = it.get("itemId")
            lbl = it.get("websiteLabel")
            if iid and lbl:
                specs.append(KapItemSpec(item_id=str(iid), website_label=str(lbl)))
        return specs

    def _match_item_ids(self, items: List[KapItemSpec], wanted_labels: List[str]) -> Dict[str, KapItemSpec]:
        matched: Dict[str, KapItemSpec] = {}
        for lab in wanted_labels:
            target = lab.lower()
            for it in items:
                if target in it.website_label.lower():
                    matched[lab] = it
                    break
        return matched

    def _parse_period_type(self, periyot: Any) -> str:
        p = str(periyot).strip()
        if p == "1":
            return PeriodTypeEnum.Q1
        if p == "2":
            return PeriodTypeEnum.Q2
        if p == "3":
            return PeriodTypeEnum.Q3
        if p == "4":
            return PeriodTypeEnum.Q4
        # KAP uses 03/06/09/12 style.
        if p == "03":
            return PeriodTypeEnum.Q1
        if p == "06":
            return PeriodTypeEnum.Q2
        if p == "09":
            return PeriodTypeEnum.Q3
        if p == "12":
            return PeriodTypeEnum.ANNUAL
        # fallback
        return PeriodTypeEnum.ANNUAL

    def _parse_currency_and_scale(self, raw: Any) -> Tuple[str, int]:
        s = ("" if raw is None else str(raw)).strip().upper()
        # Examples seen: 1000TL
        if "USD" in s:
            return CurrencyEnum.USD, 1000 if s.startswith("1000") else 1
        return CurrencyEnum.TL, 1000 if s.startswith("1000") else 1

    def _parse_numeric(self, raw: Any) -> Optional[float]:
        if raw is None:
            return None
        if isinstance(raw, (int, float)):
            return float(raw)
        s = str(raw).strip()
        if not s:
            return None
        # KAP exports may use dot as thousands or already numeric; handle both.
        s = s.replace(".", "").replace(",", ".")
        s = re.sub(r"[^0-9.\-]", "", s)
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    def _read_xlsx_table(self, xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]

        header_row_idx: Optional[int] = None
        headers: List[str] = []

        # Find header row where first column is "Şirket"
        for r in range(1, min(ws.max_row, 60) + 1):
            v = ws.cell(r, 1).value
            if v and str(v).strip() == "Şirket":
                header_row_idx = r
                for c in range(1, ws.max_column + 1):
                    hv = ws.cell(r, c).value
                    if hv is None:
                        headers.append("")
                    else:
                        headers.append(str(hv).strip())
                break

        if not header_row_idx:
            raise ValueError("KAP export xlsx: could not locate header row")

        rows: List[Dict[str, Any]] = []
        for r in range(header_row_idx + 1, ws.max_row + 1):
            values = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
            if not any(v is not None and str(v).strip() != "" for v in values):
                continue
            row: Dict[str, Any] = {}
            for h, v in zip(headers, values):
                if not h:
                    continue
                row[h] = v
            rows.append(row)

        if not rows:
            raise ValueError(
                "KAP export xlsx contains no data rows (header-only). "
                "This usually indicates that the request payload does not match KAP UI selections."
            )
        return rows

    def _get_or_create_period(
        self,
        db: Session,
        company: Company,
        year: int,
        period_type: str,
        report_type: str,
        currency: str,
    ) -> FinancialPeriod:
        period = (
            db.query(FinancialPeriod)
            .filter(
                FinancialPeriod.company_id == company.id,
                FinancialPeriod.year == year,
                FinancialPeriod.period_type == period_type,
                FinancialPeriod.report_type == report_type,
                FinancialPeriod.currency == currency,
            )
            .first()
        )
        if period:
            return period

        period = FinancialPeriod(
            company_id=company.id,
            year=year,
            period_type=period_type,
            report_type=report_type,
            currency=currency,
        )
        db.add(period)
        db.commit()
        db.refresh(period)
        return period

    def _get_or_create_account(
        self,
        db: Session,
        name_tr: str,
        statement_type: str,
        category: str,
        parent: Optional[FinancialAccount],
        level: int,
        display_order: int,
    ) -> FinancialAccount:
        q = db.query(FinancialAccount).filter(
            FinancialAccount.name_tr == name_tr,
            FinancialAccount.statement_type == statement_type,
        )
        acc = q.first()
        if acc:
            updated = False
            if acc.parent_id != (parent.id if parent else None):
                acc.parent_id = parent.id if parent else None
                updated = True
            if acc.level != level:
                acc.level = level
                updated = True
            if acc.display_order != display_order:
                acc.display_order = display_order
                updated = True
            if updated:
                db.commit()
            return acc

        acc = FinancialAccount(
            name_tr=name_tr,
            statement_type=statement_type,
            category=category,
            parent_id=parent.id if parent else None,
            level=level,
            display_order=display_order,
            is_subtotal=False,
        )
        db.add(acc)
        db.commit()
        db.refresh(acc)
        return acc

    def _get_or_create_source_map(
        self,
        db: Session,
        statement_type: str,
        raw_name: str,
        parent_raw_name: Optional[str],
        account: FinancialAccount,
    ) -> None:
        existing = (
            db.query(FinancialAccountSourceMap)
            .filter(
                FinancialAccountSourceMap.provider == "kap",
                FinancialAccountSourceMap.statement_type == statement_type,
                FinancialAccountSourceMap.raw_name == raw_name,
                FinancialAccountSourceMap.parent_raw_name == parent_raw_name,
            )
            .first()
        )
        if existing:
            if existing.account_id != account.id:
                existing.account_id = account.id
                db.commit()
            return
        mapping = FinancialAccountSourceMap(
            provider="kap",
            statement_type=statement_type,
            raw_name=raw_name,
            parent_raw_name=parent_raw_name,
            account_id=account.id,
        )
        db.add(mapping)
        db.commit()

    def fetch_and_save_financial_summary(
        self,
        db: Session,
        ticker: str,
        year_pairs: List[Tuple[int, str]],
        period: str,
    ) -> bool:
        """Fetch KAP summary items for at most 2 years per call.

        year_pairs: list of (year, period_str) where period_str is one of 03/06/09/12.
        period: kept for compatibility; not used.
        """
        try:
            company = db.query(Company).filter(Company.ticker == ticker.upper()).first()
            if not company:
                logger.error("Company %s not found in DB", ticker)
                return False

            member_oid, title, company_type = self._find_member_oid_title_and_type(ticker)

            compare_items = self._get_compare_items()

            # Minimum viable set: balance + income summary columns.
            wanted_labels = [
                "Toplam Varlıklar",
                "Dönen Varlıklar",
                "Duran Varlıklar",
                "Kısa Vadeli Yükümlülükler",
                "Uzun Vadeli Yükümlülükler",
                "Toplam Özkaynaklar",
                "Ana Ortaklığa Ait Özkaynaklar",
                "Hasılat",
                "Net Dönem Kârı",
                "Brüt Kâr",
                "Esas Faaliyet Kârı",
            ]
            matched = self._match_item_ids(compare_items, wanted_labels)
            item_ids = [spec.item_id for spec in matched.values()]

            # KAP UI limits years to max 2. If called with >2, caller must batch.
            years = [str(y) for y, _ in year_pairs]
            if len(set(years)) > 2:
                raise ValueError("KAP export supports max 2 years per request")

            periods = sorted({str(p).strip() for _, p in year_pairs})

            payload = {
                "companyType": company_type,
                "mkkMemberIdList": [member_oid],
                "mkkMemberTitleList": [title],
                "yearList": sorted({str(y) for y, _ in year_pairs}),
                "periodList": periods,
                "itemIdList": item_ids,
                "sectors": [self.sector],
            }

            xlsx_bytes, ctype = self._post_bytes(self.EXPORT_COMPARE_ITEMS_URL, payload)
            if "spreadsheetml" not in (ctype or ""):
                logger.warning("Unexpected KAP export content-type: %s", ctype)

            rows = self._read_xlsx_table(xlsx_bytes)

            # Prepare accounts hierarchy (minimal)
            # Balance sheet root/group
            bs_root = self._get_or_create_account(
                db, "KAP Özet Bilanço", StatementTypeEnum.BALANCE_SHEET, "asset", None, 0, 0
            )
            income_root = self._get_or_create_account(
                db, "KAP Özet Gelir Tablosu", StatementTypeEnum.INCOME_STATEMENT, "revenue", None, 0, 0
            )

            order = 1
            account_by_header: Dict[str, FinancialAccount] = {}

            def ensure_account(header: str) -> FinancialAccount:
                nonlocal order
                cleaned = str(header).strip()
                if cleaned in account_by_header:
                    return account_by_header[cleaned]
                stmt, cat, parent = self._classify_header(header, bs_root, income_root)
                acc = self._get_or_create_account(
                    db, header, stmt, cat, parent, 1, order
                )
                self._get_or_create_source_map(db, stmt, header, parent.name_tr if parent else None, acc)
                account_by_header[cleaned] = acc
                order += 1
                return acc

            # SessionLocal is configured with autoflush=False in this project.
            # KAP export can include multiple rows for the same period (different announcement IDs),
            # so we must keep a per-period upsert cache across all rows.
            values_cache_by_period_id: Dict[int, Dict[int, FinancialValue]] = {}

            for row in rows:
                year_raw = row.get("Yıl")
                per_raw = row.get("Periyot")
                if year_raw is None or per_raw is None:
                    continue
                year = int(str(year_raw).strip())
                # KAP period can be quarter number (1-4) or month-style (03/06/09/12).
                # Don't zfill quarter values ("4" -> "04"), it would break _parse_period_type.
                per = str(per_raw).strip()
                if per.endswith(".0"):
                    per = per[:-2]

                currency_raw = row.get("Sunum Para Birimi")
                currency, scale = self._parse_currency_and_scale(currency_raw)

                period_type = self._parse_period_type(per)
                # KAP export provides "Finansal Tablo Niteliği" (Konsolide, etc.)
                report_raw = str(row.get("Finansal Tablo Niteliği") or "konsolide").strip().lower()
                report_type = ReportTypeEnum.KONSOLIDE if "kons" in report_raw else ReportTypeEnum.STANDART

                period_obj = self._get_or_create_period(db, company, year, period_type, report_type, currency)

                value_by_account_id = values_cache_by_period_id.get(period_obj.id)
                if value_by_account_id is None:
                    value_by_account_id = {
                        fv.account_id: fv
                        for fv in db.query(FinancialValue)
                        .filter(FinancialValue.period_id == period_obj.id)
                        .all()
                    }
                    values_cache_by_period_id[period_obj.id] = value_by_account_id

                for header, val in row.items():
                    if header in {
                        "Şirket",
                        "Bildirim ID",
                        "Bildirim Yayın Tarihi",
                        "Yıl",
                        "Periyot",
                        "Finansal Tablo Niteliği",
                        "Sunum Para Birimi",
                        "Sektörel Tablo Türü",
                    }:
                        continue
                    acc = ensure_account(header)
                    num = self._parse_numeric(val)
                    if num is None:
                        continue
                    num = num * scale

                    existing = value_by_account_id.get(acc.id)
                    if existing is not None:
                        existing.value = num
                        continue

                    new_val = FinancialValue(period_id=period_obj.id, account_id=acc.id, value=num)
                    db.add(new_val)
                    value_by_account_id[acc.id] = new_val

            db.commit()
            return True
        except Exception:
            logger.exception("KAP financial summary fetch failed for %s", ticker)
            db.rollback()
            return False

    def _classify_header(
        self,
        header: str,
        bs_root: FinancialAccount,
        income_root: FinancialAccount,
    ) -> Tuple[str, str, FinancialAccount]:
        h = header.upper()
        # Balance sheet heuristics
        if any(k in h for k in ["VARLIK", "YÜKÜMLÜ", "ÖZKAYNAK"]):
            # Keep everything under balance sheet root
            if "ÖZKAYNAK" in h:
                return StatementTypeEnum.BALANCE_SHEET, "equity", bs_root
            if "YÜKÜML" in h or "BORÇ" in h:
                return StatementTypeEnum.BALANCE_SHEET, "liability", bs_root
            return StatementTypeEnum.BALANCE_SHEET, "asset", bs_root

        # Income statement heuristics
        if any(k in h for k in ["HASILAT", "KÂR", "KAR", "ZARAR", "SATIŞ"]):
            if "HASILAT" in h or "SATIŞ" in h:
                return StatementTypeEnum.INCOME_STATEMENT, "revenue", income_root
            return StatementTypeEnum.INCOME_STATEMENT, "expense", income_root

        # Fallback: treat as balance sheet metric.
        return StatementTypeEnum.BALANCE_SHEET, "asset", bs_root
